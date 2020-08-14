import openpyxl

def read_excel(file):
    """
    得到一个三维数组，第一维sheet name
    [
	['Sheet', [
		['href', 'zhongbiaoren', 'xmjl', 'ggname', 'zbtime', 'zbsl_count'],
		['https://hhb.cbi360.net/sg_1360545/jzs/2330804.html', '普洱建一建筑有限责任公司', '张勇', '澜沧县2019年度农田建设项目第四标段', '2020-05-27', 1],
		['https://hhb.cbi360.net/sg_1281288/jzs/2330831.html', '普洱品筑建设有限责任公司', '张魏忠', '澜沧县2019年度农田建设项目第三标段', '2020-05-27', 1]
	]]
]
    :param file:
    :return:
    """
    wb = openpyxl.load_workbook(file)
    sheetNames = wb.sheetnames
    all_sheet_data=[]
    for sheetName in sheetNames:
        ws = wb[sheetName]
        max_row = ws.max_row
        max_col = ws.max_column
        data=[]
        for row in range(1,max_row+1):
            col_value=[]
            for col in range(1,max_col+1):
                cell_data=ws.cell(row=row,column=col).value
                col_value.append(cell_data)
            data.append(col_value)
        all_sheet_value_tmp=[sheetName,data]
        all_sheet_data.append(all_sheet_value_tmp)
    return all_sheet_data


if   __name__=='__main__':
    infilename = r"D:\SVN\业务数据维护\项目经理业绩_胡金花_20200603_153200.xlsx"
    all_sheet_data = read_excel(infilename)
    print(all_sheet_data[0][1][1:])

    # for sheetConte in all_sheet_data:
    #     # 得到sheet名字
    #     sheetName = sheetConte[0]
    #     print(sheetName)
    #
    #     # 得到除了第一行(字段名字)的所有表数据
    #     indata = sheetConte[1][1:]
    #     print(indata)
    #
    #     for row in indata:
    #         zhongbiaoren = row[1]
    #         print(zhongbiaoren)
    #         xmjl = row[2]
    #         print(xmjl)






