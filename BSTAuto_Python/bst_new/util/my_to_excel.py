# -*- coding: utf-8 -*-
import os
import openpyxl
import csv
import pandas as pd

def  wirteDataToExcel(outfilename,sheetName,columnRows, recordRows):
    """
    文件存在就不能再写入了
    :param outfilename:
    :param sheetName:
    :param columnRows:
    :param recordRows:
    :return:
    """
    if os.path.exists(outfilename):
        raise FileExistsError("file exists")

    wb = openpyxl.Workbook()
    # wb.create_sheet(sheetName,0)
    # sheet =wb.get_sheet_by_name(sheetName)
    sheet=wb.active

    # 当列数超过26时，添加AA,AB,AC...
    myAlphbet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                 'U', 'V', 'W', 'X', 'Y', 'Z']
    # 第一行写入数据库表列名:一维字符串数组
    for i in range(len(columnRows)):
        loc = myAlphbet[i]+str(1)
        sheet[loc]=columnRows[i]

    #写入数据库表所有数据，二维字符串数组
    for row in  range(len(recordRows)):
        sheet.append(recordRows[row])
        # for col in range(len(recordRows[row])):
        #     loc = myAlphbet[col]+str(row+2)
        #     sheet[loc] = recordRows[row][col]

    wb.save(outfilename)


def  wirteDataToExcel2(outfilename,headers, recordRows,headers_exists=0):
    """
    文件存在可以在后面追加内容
    :param outfilename:
    :param headers:
    :param recordRows:
    :param headers_exists:
    :return:
    """
    if os.path.exists(outfilename):
        # raise FileExistsError("file exists")
        wb = openpyxl.load_workbook(outfilename)
    else:
        wb = openpyxl.Workbook()
        # wb.create_sheet(sheetName,0)
        # sheet =wb.get_sheet_by_name(sheetName)

    sheet=wb.active
    # 当列数超过26时，添加AA,AB,AC...
    myAlphbet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                 'U', 'V', 'W', 'X', 'Y', 'Z']
    # 第一行写入数据库表列名:一维字符串数组,只写入一次
    if headers_exists == 0:
        for i in range(len(headers)):
            loc = myAlphbet[i]+str(1)
            sheet[loc]=headers[i]

    #写入数据库表所有数据，二维字符串数组
    for row in  range(len(recordRows)):
        sheet.append(recordRows[row])
        # for col in range(len(recordRows[row])):
        #     loc = myAlphbet[col]+str(row+2)
        #     sheet[loc] = recordRows[row][col]

    wb.save(outfilename)

def  writeToCsv():
    with open("my.csv", "a", newline='') as f:
        writer = csv.writer(f)
        writer.writerow(["URL", "predict", "score"])
        row = [['1', 1, 1], ['2', 2, 2], ['3', 3, 3]]
        for r in row:
            writer.writerow(r)


def  pdwriteToExcel():
    result_list = [['1', 1, 1], ['2', 2, 2], ['3', 3, 3]]
    columns = ["URL", "predict", "score"]
    dt = pd.DataFrame(result_list, columns=columns)
    dt.to_excel("result_xlsx.xlsx", index=0)
    dt.to_csv("result_csv.csv", index=0)


def test02():
    print('success')